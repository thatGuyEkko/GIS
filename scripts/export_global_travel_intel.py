from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "data" / "global-travel-intel-template.xlsx"
OUTPUT_PATH = ROOT / "data" / "global-travel-intel.json"
SHEET_NAMES = ["countries", "news", "opportunities", "expos"]


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    payload = {
        "metadata": {
            "source_workbook": WORKBOOK_PATH.name,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "sheet_names": SHEET_NAMES,
        }
    }

    for sheet_name in SHEET_NAMES:
        payload[sheet_name] = read_sheet(workbook[sheet_name])

    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Exported {OUTPUT_PATH.relative_to(ROOT)} from {WORKBOOK_PATH.relative_to(ROOT)}")


def read_sheet(sheet) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]
    results: list[dict[str, Any]] = []

    for values in rows[1:]:
        if is_empty_row(values):
            continue

        item: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue

            item[header] = normalize_value(values[index] if index < len(values) else None)

        results.append(item)

    return results


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(value in (None, "") for value in values)


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, bool):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value

    return value


if __name__ == "__main__":
    main()
